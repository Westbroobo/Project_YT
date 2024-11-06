"""
Filename:ebay_vehicle.py
Author: Westbroobo
Date: 2024/9/29
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import requests
from tool import UA, Proxy
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'Item', 'Vehicle', 'Engine', 'Vehicle_Engine'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 65
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
wb.save('eBay_vehicle.xlsx')

df_menu = pd.read_excel('./menu.xlsx',header=0)
items = df_menu['Item'].tolist()

def list_cleand(lst):
    new_list = []
    lst.sort()
    for x, y in zip(lst, lst[1:]):
        if int(y) - int(x) > 1:
            new_list.append(lst[:lst.index(y)])
            lst = lst[lst.index(y):]
    new_list.append(lst)
    return new_list

def fetch_vehicle(rows, list_table_title, make_model_year_dict, year_make_model_engine_ls, engine_ls):
    try:
        index_year = list_table_title.index('Year')
        index_make = list_table_title.index('Make')
        index_model = list_table_title.index('Model')
        index_engine = list_table_title.index('Engine')
        for row in rows:
            cells = row["cells"]
            year = int(cells[index_year]["textSpans"][0]["text"])
            make = cells[index_make]["textSpans"][0]["text"]
            model = cells[index_model]["textSpans"][0]["text"]
            engine = cells[index_engine]["textSpans"][0]["text"]
            year_make_model_engine = str(year) + ' ' + make + ' ' + model + ' ' + engine
            if year_make_model_engine not in year_make_model_engine_ls:
                year_make_model_engine_ls.append(year_make_model_engine)
            if 'Trim' in list_table_title and 'L' in engine:
                engine_ls.append(engine.split(' ')[0])
            if 'Variant' in list_table_title and 'cc' in engine:
                engine_ls.append(engine.split(' ')[0])
            if make not in make_model_year_dict:
                make_model_year_dict[make] = {model: [year]}
            else:
                if model not in make_model_year_dict[make]:
                    make_model_year_dict[make][model] = [year]
                else:
                    make_model_year_dict[make][model].append(year)
            make_model_year_dict[make][model] = list(set(make_model_year_dict[make][model]))
        return make_model_year_dict, year_make_model_engine_ls, engine_ls
    except:
        index_year = list_table_title.index('Baujahr')
        index_make = list_table_title.index('Marke')
        index_model = list_table_title.index('Modell')
        index_engine = list_table_title.index('Typ')
        for row in rows:
            cells = row["cells"]
            year = cells[index_year]["textSpans"][0]["text"]
            make = cells[index_make]["textSpans"][0]["text"]
            model = cells[index_model]["textSpans"][0]["text"]
            engine = cells[index_engine]["textSpans"][0]["text"]
            year_make_model_engine = make + ' ' + model + ' ' + year + ' ' + engine
            if year_make_model_engine not in year_make_model_engine_ls:
                year_make_model_engine_ls.append(year_make_model_engine)
            engine_ls.append(engine)
            year_ = year.split('-')
            year_1 = year_[0][:4]
            year_2 = year_[-1][:4]
            year__ = [y for y in range(int(year_1), int(year_2)+1)]
            if make not in make_model_year_dict:
                make_model_year_dict[make] = {model: year__}
            else:
                if model not in make_model_year_dict[make]:
                    make_model_year_dict[make][model] = year__
                else:
                    for y_ in year__:
                        make_model_year_dict[make][model].append(y_)
            make_model_year_dict[make][model] = list(set(make_model_year_dict[make][model]))
        return make_model_year_dict, year_make_model_engine_ls, engine_ls

def dict_to_str(make_model_year_dict):
    make_model_year_ls = []
    for make in make_model_year_dict:
        for model in make_model_year_dict[make]:
            if len(make_model_year_dict[make][model]) == 1:
                content = make + ' ' + model + ' ' + str(make_model_year_dict[make][model][0])
                make_model_year_ls.append(content)
            elif int(max(make_model_year_dict[make][model])) - int(min(make_model_year_dict[make][model])) == len(
                    make_model_year_dict[make][model]) - 1:
                content = make + ' ' + model + ' ' + str(min(make_model_year_dict[make][model])) \
                          + '-' + str(max(make_model_year_dict[make][model]))
                make_model_year_ls.append(content)
            else:
                new_list = list_cleand(make_model_year_dict[make][model])
                for n in new_list:
                    if len(n) == 1:
                        content = make + ' ' + model + ' ' + str(n[0])
                        make_model_year_ls.append(content)
                    else:
                        content = make + ' ' + model + ' ' + str(min(n)) + '-' + str(max(n))
                        make_model_year_ls.append(content)
    make_model_year_ls = list(set(make_model_year_ls))
    make_model_year_ls.sort()
    make_model_year = '\n'.join(make_model_year_ls)
    return make_model_year

cookies = {
    '__uzme': '5209',
    'cid': 'rBiPWaMlw0K7sOkA%23469222263',
    '__ssds': '2',
    '_gcl_au': '1.1.1193484553.1723596701',
    '_fbp': 'fb.1.1723596701882.643149592300209675',
    '__gsas': 'ID=151fc2379bf4647b:T=1723597742:RT=1723597742:S=ALNI_MbFXDC6BZ7uSb0mEzNmu72Sp4WLXg',
    'AMP_MKTG_f93443b04c': 'JTdCJTdE',
    '__uzma': 'df81ef4a-4a77-4888-bb6f-1a6793f891b3',
    '__uzmb': '1730251461',
    '__uzmaj2': '2dfea98a-f162-4646-bfb2-5e036735d06d',
    '__uzmbj2': '1730251462',
    'utag_main__sn': '2',
    '_pin_unauth': 'dWlkPU1EVTJaR1ZoT1RndE9ETXdZaTAwWmpSaUxXSXdNRGt0WVRnMk1tTXlaRFJqTlRZNA',
    '_scid': '1VkDPcDhylu-s5VIJen-IQJuz7eoJUET',
    '_ScCbts': '%5B%5D',
    '_scid_r': '4NkDPcDhylu-s5VIJen-IQJuz7eoJUET4NELYg',
    '__uzmlj2': 'FBZdcKGkbWPjwhygHTkO2yI1X8TRDw3AbwMTOWm6G6I=',
    'shs': 'BAQAAAZK70qguAAaAAVUAD2kCvpgyMjE4NjUxOTQxMDAzLDIzb1YqFj3egvFcE1TEcx4+MAh2Kg**',
    '__gads': 'ID=124e20675a7fb8f2:T=1730421319:RT=1730421319:S=ALNI_MYAjkbXqrxJxTL9fV4xNbfQnDi-3g',
    '__gpi': 'UID=00000f495bd018ba:T=1730421319:RT=1730421319:S=ALNI_MYnxMQu861UQKLIDlhC-gPO9uMKBg',
    '__eoi': 'ID=82f40705557af401:T=1730421319:RT=1730421319:S=AA-AfjY3R1DqafEJ1VZx8ziZtQWq',
    '__ssuzjsr2': 'a9be0cd8e',
    'ak_bmsc': 'F69E18AB35E953C121E5A1BB6846DBC4~000000000000000000000000000000~YAAQlAQsF3oKgs2SAQAAJU485hkJMKCfV/bi0Q+k2VvqpJT+bbFkYBYLr9l82FFgarHpaRBSiCiE4bapnTPY6jA+OfU8hKJEna9E/bKkBLl8V3XEI3XsyZWCi1Go047BVVl7fdoQw1bg1mLPFkOENKwRjySstItj+go5kB5aoVhw/5uy7yPwX1vv4f3FvBTR5XfSkhM1+Ky7hFcjXZ24J0aokppyGfcV8Wn6oYol4adJbZhvFT835kvdHQgP93uuvN59WUk+I7tuFwmnD17+LKAxdBaNFzWfxgUBeRcsphznK2RSNCnUIecS6BQkM3cVNf3jNu1NCwq4C9WunUVE4hf7KNFUoH4pAxWn2XJ3kAWHM5T2MyoVu9RbrfL/xawmec+RXuKIWfJkmA==',
    '__uzmcj2': '52479106617908',
    '__uzmdj2': '1730442511',
    '__uzmfj2': '7f6000e86307bf-6fc9-4397-8068-4a0a2f5ef0a61730251462620191049282-5c012ec4907aea261066',
    's': 'CgADuAF5nJcqYMwZodHRwczovL3d3dy5lYmF5LmNvbS9zY2gvaS5odG1sP19mcm9tPVI0MCZfbmt3PU1EMzM0MDM5Jl9zYWNhdD0wJl9zb3A9MTUmTEhfUHJlZkxvYz0xJnJ0PW5jBwD4ACBnJccLZGIwM2Q4YjAxOTIwYTU2MDE1ZmUzMjZmZmE1MTlhNzZ4b1jJ',
    '__uzmc': '19661167570741',
    '__uzmd': '1730443545',
    '__uzmf': '7f6000e86307bf-6fc9-4397-8068-4a0a2f5ef0a61730251461328192084388-d3a4bfa57a33cc3a1675',
    'AMP_f93443b04c': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjIyNWQ1NjBiZi05YWVjLTRjMTYtYWIxZi0wMzk2ODJjNDAwMzYlMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzMwNDM5MDYxMTM2JTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTczMDQ0MzU0Nzc0NSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMjEwNCUyQyUyMnBhZ2VDb3VudGVyJTIyJTNBMjM3JTdE',
    '__deba': 'PvJrqtq_rHI8m_-l6zyRJPTVcD-bB05e1zogDigSDXjWZK5gDoilvetJ5eQme8VA0Mv9jKWJoVy08mDLSPFacV__S2OG1PGjDtaXqL7j2dYcs6ZJt46kRLNWkfYNJ9FX_MBDYH0Fdw96NC47ausaVw==',
    'bm_sv': '0D94CFD71A219AF751C31661341B0969~YAAQGtU+F3mHfOOSAQAAAIN55hmjLLqboNMVqdKOPuh2dO703zCZeAYL8vCalkdMgZc02pTwbxqV8xg4MuHf46ZNJmGcBJYqcwnUvVCLg2QasyFz2K0KkIwKEpI89zrafwixApiTKUn+VGKQYTSfmDuGqUcshzKn4dwqTBVR+zhKLf1xcvFWYrpuh9EcxubHuseleMEJ/sjzM6aKcQSSZjPteRmsKRgtl7Fu2Cw+J+qcT1GW1NaGuAkNCIfQ/reD~1',
    'ns1': 'BAQAAAZJU03gwAAaAAKUADWkFrLwyMzYyODgzMzM1LzA7ANgAU2kFrLxjNjl8NjAxXjE3MzAyNTE0NjEzOTdeXjFeM3wyfDV8NHw3fDEwfDQyfDQzfDExXl5eNF4zXjEyXjEyXjJeMV4xXjBeMV4wXjFeNjQ0MjQ1OTA3NYV8G4lUZy357a/YViPRaZUJhLfn',
    'dp1': 'bu1p/eWFuZy0xMDE0OTQ*6ae6e03c^kms/in6ae6e03c^pbf/%230000e000e000000080800000006905acbc^u1f/yangteng6ae6e03c^bl/USen-US6ae6e03c^',
    'nonsession': 'BAQAAAZJU03gwAAaAAAQAC2kCvph5YW5nLTEwMTQ5NAAQAAtpBay8eWFuZy0xMDE0OTQAMwAOaQWsvDkxNzEwLTU0MTcsVVNBAEAAC2kFrLx5YW5nLTEwMTQ5NACcADhpBay8blkrc0haMlByQm1kajZ3Vm5ZK3NFWjJQckEyZGo2TUhsb0NwQUpPQm9nbWRqNng5blkrc2VRPT0AnQAIaQWsvDAwMDAwMDAxAMoAIGrm4DxkYjAzZDhiMDE5MjBhNTYwMTVmZTMyNmZmYTUxOWE3NgDLAANnJIBEMzUwAWQAB2rm4DwjMDAwMDBhG9/hJf+6aniiF0NNUPxbBTw1FLE*',
    'totp': '1730443579899.pmgwYgnbMVwNdZD7V5wfJn2V2cHzWQNeAUwuU2IVYbCnhukKxAejhmMRPBlIyZu08jfmv0OWyT5qXS+VDi4tnQ==.9pE0cMiYgg7ASKb6GGgm8BXYNCnNT14gmViydu8hUCU',
    'ds2': 'sotr/bGvQkzzzzzzz^',
    'ebay': '%5Ejs%3D1%5Esbf%3D%23000000%5Epsi%3DAeQS1B9c*%5E',
}

def fetch_data(url, item):
    list_table_title = []
    make_model_year_dict = {}
    year_make_model_engine_ls = []
    engine_ls = []
    headers = {
        'accept': 'application/json',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cache-control': 'max-age=0',
        'content-type': 'application/json',
        'origin': 'https://www.ebay.com',
        'priority': 'u=1, i',
        'referer': url,
        'sec-ch-ua': '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
        'sec-ch-ua-full-version': '"128.0.6613.113"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-model': '""',
        'sec-ch-ua-platform': '"Windows"',
        'sec-ch-ua-platform-version': '"10.0.0"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': UA.get_UA(),
        'x-ebay-answers-request': 'pageci=8150e95c-7213-4f7a-98bb-585598255d75,parentrq=b16c28a01910ab7e0f393b0dffebf16b',
        'x-ebay-c-correlation-session': 'operationId=4429486',
        'x-ebay-c-tracking-config': 'viewTrackingEnabled=true,perfTrackingEnabled=true,navTrackingEnabled=true,navsrcTrackingEnabled=true,swipeTrackingEnabled=false,showDialogTrackingEnabled=true',
    }
    json_data = {
            'scopedContext': {
                'catalogDetails': {
                    'itemId': str(item),
                    'categoryId': '33602',
                    'marketplaceId': 'EBAY-US',
                },
            },
        }
    session = requests.session()
    for i in range(0, 10000, 20):
        num = 100
        e_ = ''
        for test in range(num):
            params = {
                'module_groups': 'PART_FINDER',
                'referrer': 'VIEWITEM',
                'offset': str(i),
                'module': 'COMPATIBILITY_TABLE',
            }
            try:
                rsp = session.post('https://www.ebay.com/g/api/finders', params=params, cookies=cookies,
                                         headers=headers, json=json_data, proxies=Proxy.get_Proxy_Requests())
                status_code = rsp.status_code
                data = rsp.json()
                if i == 0:
                    titles = data["modules"]["COMPATIBILITY_TABLE"]["paginatedTable"]["header"]["cells"]
                    for title in titles:
                        list_table_title.append(title["textSpans"][0]["text"])
                rows = data["modules"]["COMPATIBILITY_TABLE"]["paginatedTable"]["rows"]
                if status_code == 200:
                    break
            except Exception as e:
                if type(e) == requests.exceptions.JSONDecodeError:
                    e_ = e
                    break
                time.sleep(0.3)
                continue
        if type(e_) == requests.exceptions.JSONDecodeError:
            session.close()
            break
        if test == num - 1:
            print('{}【尝试次数：{}】-error'.format(item, test+1))
            break
        if len(rows) == 0:
            session.close()
            break
        make_model_year_dict, year_make_model_engine_ls, engine_ls = fetch_vehicle(rows, list_table_title,
                                                                                   make_model_year_dict,
                                                                                   year_make_model_engine_ls, engine_ls)
    if test != num - 1:
        vehicle = dict_to_str(make_model_year_dict)
        vehicle_engine = '\n'.join(year_make_model_engine_ls).strip()
        engine_ls = list(set(engine_ls))
        engine_ls.sort()
        engine_all = '\n'.join(engine_ls)
        wb = load_workbook(filename='./eBay_vehicle.xlsx')
        ws = wb['Data']
        serial = ws.max_row
        ws.append([serial, str(item), vehicle, engine_all, vehicle_engine])
        last_row = ws.max_row
        ws.row_dimensions[last_row].height = 20
        for col in range(1, 6):
            cell = ws.cell(row=last_row, column=col)
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        wb.save(filename='./eBay_vehicle.xlsx')
        wb.close()
        print('{}-success，进度：{}/{}'.format(item, serial, len(items)))

pool = pool.Pool(15)
for item in items:
    url = 'https://www.ebay.com/itm/' + str(item)
    pool.spawn(fetch_data, url, item)
pool.join()


