"""
Filename:2.ebay_US.py
Author: Westbroobo
Date: 2024/8/30
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import pandas as pd
import requests
from tool import UA, Proxy
from lxml import etree
import json
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'Item', 'Title', 'Img', 'Seller', 'Price', 'Availability', 'Sold', 'Json_specifics', 'Item_description', 'Url'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 45
ws.column_dimensions['J'].width = 28
ws.column_dimensions['K'].width = 28
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
wb.save('eBay.xlsx')

cookies = {
                '__uzma': '8efa63b2-5b59-482c-b4b5-9487d7862a23',
                '__uzmb': '1723596676',
                '__uzme': '5209',
                'AMP_MKTG_f93443b04c': 'JTdCJTdE',
                'cid': 'rBiPWaMlw0K7sOkA%23469222263',
                '__ssds': '2',
                '__uzmaj2': '9ed18687-3fb2-466d-a0db-180c7e6eca5e',
                '__uzmbj2': '1723596687',
                'utag_main__sn': '1',
                '_gcl_au': '1.1.1193484553.1723596701',
                '_fbp': 'fb.1.1723596701882.643149592300209675',
                '__uzmlj2': 'bJVcz5GtV5oJp3toDqqWMSPwdcPlcLIHD8QlMoEa3bg=',
                '__gsas': 'ID=151fc2379bf4647b:T=1723597742:RT=1723597742:S=ALNI_MbFXDC6BZ7uSb0mEzNmu72Sp4WLXg',
                'shs': 'BAQAAAZGG1Rg0AAaAAVUAD2idMzkyMjE4NjUxOTQxMDAzLDKN9yHc8y2CBeYlJzVICPkpT7TxnQ**',
                '__ssuzjsr2': 'a9be2cd8e',
                '__uzmcj2': '86533135768362',
                '__uzmdj2': '1725005787',
                '__uzmfj2': '7f60005a929bdc-3b5b-45e2-ab0b-761e825af44f17235966878211409099289-a843813f13acece91354',
                '__gads': 'ID=8cd97c1e52c80d01:T=1725001318:RT=1725005895:S=ALNI_MYpF9ZolBeqKKzNiWci2T07KAGvIQ',
                '__gpi': 'UID=00000ee32c1e45ae:T=1725001318:RT=1725005895:S=ALNI_MYM8h0OXm4Jt16LhrRZ5jjV2p3NRg',
                '__eoi': 'ID=3941d8cfede7ca0b:T=1725001318:RT=1725005895:S=AA-AfjZtdd1vn3IPZZl_ynP5K3Pw',
                's': 'CgADuABhm1n9GMAZodHRwczovL3d3dy5lYmF5LmNvbS8HAPgAIGbWXBw0ZTVlMTM1ZTE5MTBhYWIzNmRjNTZmZThmZmZlM2M1N9nBn2M*',
                'ak_bmsc': '74A8D75BFAC6739116175352586A0F8C~000000000000000000000000000000~YAAQBopFyz2Azo6RAQAAQTRRsRgSn/py04k2NVoBWErusHb+mOIjH1FEr11+HhYlr0vXeg5rB2jOXmFEY1vFrhfEMp6Fl4GUNDO6yiY4xBeRendpJWEm/BoxCpz4QXEaa1NYzlB6Pmd7qfZTmIeVGjX8its6aPesgkwiOBR18kr3bLkHsNDiy2jhrLuE0NiwzQSNdIms8h6brsbvRu0gZuyaKpzpXHdW0fSJaOCNo/AV7RJq+luR30ddkXsff4M6NVXQv+a7L4X2C/9ES8Qc/IL7hkGa45qNQnGYsvRwjftmjosh/ic8Sz5hD5gwtTIYpzoUBUhNaejz2F2bN/FPAqLyZpsiUFVTto5YzZ3KMKNRDdh3D2/Hwkc8ONhjymcvZn/esabwyrs=',
                '__uzmd': '1725256776',
                '__deba': 'BSMonH9oZbLeqKowEqQYz2gd1XX5v01wYGfwEkKFlDC2ehdlQBztvF0zt4JwzAcHI2i1hbzWKfbBPKjzqLTdRLa7ci_qb88uyoNVlijHlw3W6v-QdEo83d5arzxSi_HlbkZP38YN2ObMqXv073kLfw==',
                '__uzmc': '62415247090334',
                '__uzmf': '7f60005a929bdc-3b5b-45e2-ab0b-761e825af44f17235966765551660099950-ee524e0826b0b9032470',
                'ns1': 'BAQAAAZBR14g6AAaAAKUADWi2iBcyMzYyODgzMzM1LzA72VoxHqblTJgVyBCFULetRNBqx+A*',
                'dp1': 'bu1p/eWFuZy0xMDE0OTQ*6a97bb97^kms/in6a97bb97^pbf/%232000000e00020000000808000000468b68817^u1f/yangteng6a97bb97^expt/000172359675500567ac9993^bl/USen-US6a97bb97^',
                'nonsession': 'BAQAAAZBR14g6AAaAAAQAC2idMzl5YW5nLTEwMTQ5NAAQAAtotogXeWFuZy0xMDE0OTQAMwAOaLaIFzkxNzEwLTU0MTcsVVNBAEAAC2i2iBd5YW5nLTEwMTQ5NACcADhotogXblkrc0haMlByQm1kajZ3Vm5ZK3NFWjJQckEyZGo2TUhsb0NwQUpPQm9nbWRqNng5blkrc2VRPT0AnQAIaLaIFzAwMDAwMDAxAMoAIGqXu5c0ZTVlMTM1ZTE5MTBhYWIzNmRjNTZmZThmZmZlM2M1NwDLAAFm1VufNwFkAAdql7uXIzAwMDA4YW27PaWHcM+NWk8KYsdeKASeE96S',
                'bm_sv': '46926929DBE7BF475345FE528C4CF74B~YAAQBopFyxGEzo6RAQAAOnJSsRjdmuAhDgeVUiKcSgyzadLCy8sztBaTRJdQJZmVct/Xh+GOhTW6EZSESt/hpEa1qOVXVnNZ1jdOrVa5G61OpgI34H/yG1u2R2Vl21v20h+e8A/j4vPV4FtKvSKjzuDC1MnLCoiWGIxXdbCJEgw4gWCaf9JAc6keSb0NemPkeMEskB1PH0RytEIEdgxg/m1xpTuL6P/G6A1wCqnImaSD6Mm5kINlS+UamSam2Q==~1',
                'ebay': '%5Ejs%3D1%5Esbf%3D%23000000%5Epsi%3DAUTF30bc*%5E',
                'ds2': 'sotr/bGvQkzzzzzzz^',
                'totp': '1725257229796.+vjUD4YEs4Le6wAtzJAqwJ0jIUGG8e4Ky/nKuDP0KLY0U5AWFaF73+mhCYADxpdM0qV1kKHaSMW99bkWCdctig==.XMYN6I1dAkCuUSomMgov3RgCwK1fr6tOw6rXi2XYY8s',
                'AMP_f93443b04c': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJiZmFmYTk5MC1hOTNlLTRiOTEtYmUzZC01MzVhYTE2NTQ5ZWUlMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzI1MjU2MzcyNTM1JTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcyNTI1NzIzNjc5MSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMTUxNyUyQyUyMnBhZ2VDb3VudGVyJTIyJTNBMzIlN0Q=',
            }
df_menu = pd.read_excel('./menu.xlsx', header=0)
list_item = df_menu['Item'].tolist()

def fetch_data(item):
    num = 10
    for test in range(num):
        try:
            url = 'https://www.ebay.com/itm/' + str(item)
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-language': 'zh-CN,zh;q=0.9',
                'cache-control': 'max-age=0',
                'priority': 'u=0, i',
                'sec-ch-ua': '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
                'sec-ch-ua-full-version': '"128.0.6613.113"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-model': '""',
                'sec-ch-ua-platform': '"Windows"',
                'sec-ch-ua-platform-version': '"10.0.0"',
                'sec-fetch-dest': 'document',
                'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'none',
                'sec-fetch-user': '?1',
                'upgrade-insecure-requests': '1',
                'user-agent': UA.get_UA(),
                'referer': 'https://www.ebay.com/'
            }
            rsp = requests.get(url=url, cookies=cookies, headers=headers, proxies=Proxy.get_Proxy_Requests())
            # MC = re.findall('window.\\$MC\\|\\|\\[\\]\\).concat\\((.+?)</script>', rsp.text)
            status_code = rsp.status_code
            tree = etree.HTML(rsp.text)
            title = tree.xpath('//h1[@class="x-item-title__mainTitle"]/span/text()')[0]
            img = tree.xpath('//*[@class="ux-image-carousel-item image-treatment active  image"]/img')[0].get('data-zoom-src')
            while not img:
                img = tree.xpath('//*[@class="ux-image-carousel-item image-treatment active  image"]/img')[0].get('src')
            seller = tree.xpath('//*[@class="x-sellercard-atf__info__about-seller"]/a/span/text()')[0]
            price = tree.xpath('//*[@class="x-price-primary"]/span/text()')[0]
            try:
                availability = tree.xpath('//*[@class="x-quantity__availability evo"]/span[1]/text()')[0]
            except:
                availability = ''
            try:
                sold = tree.xpath('//*[@class="x-quantity__availability evo"]/span[2]/text()')[0]
            except:
                sold = ''
            try:
                Item_description = tree.xpath('//*[@id="desc_ifr"]')[0].get('src')
            except:
                Item_description = ''
            specifics_cols = tree.xpath('//*[@class="ux-layout-section-evo__col"]')
            list_ = []
            for spec_col in specifics_cols:
                dict_ = {}
                try:
                    key = spec_col.xpath('./dl/dt/div/div/span/text()')[0]
                except:
                    break
                if key == 'Condition':
                    value = spec_col.xpath('./dl/dd/div/div/span/span/span/text()')[0]
                else:
                    value = spec_col.xpath('./dl/dd/div/div/span/text()')[0]
                dict_[key] = value
                list_.append(dict_)
            formatted_data = {str(index): item for index, item in enumerate(list_)}
            Item_specifics = json.dumps(formatted_data)
            if status_code == 200:
                wb = load_workbook(filename='./eBay.xlsx')
                ws = wb['Data']
                serial = ws.max_row
                ws.append([serial, str(item), title, img, seller, price, availability, sold, Item_specifics, Item_description, url])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 15
                for col in range(1, 12):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./eBay.xlsx')
                wb.close()
                rsp.close()
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(item, test + 1, serial, len(list_item)))
                break
        except Exception as e:
            # print(e)
            time.sleep(0.3)
            if test == num - 1:
                print('{}【尝试次数：{}】-error'.format(item, test+1))
            continue


pool = pool.Pool(15)
for item in list_item:
    pool.spawn(fetch_data, item)
pool.join()


