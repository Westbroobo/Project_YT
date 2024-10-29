# -*- coding: utf-8 -*- 
# @Time ： 2024/7/13 10:20
# @Auth ： Westbroobo
# @File ：2.车型逆向.py

import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

conn = mysql.connector.connect(
    host="127.0.0.1",
    user="root",
    password="654321",
    database="data"
)
Type_Code = '1001781'
cursor = conn.cursor()
sql = '''
        SELECT CONCAT(t2.Make,' ',t2.Model) as Model_US, 
        t2.Make,
        t2.Year,
        t2.Model,
        t2.Engine,
        t2.Manufacturer,
        t2.Part_Number,
        t2.Alternate_OE_Part_Numbers,
        t2.Vehicle,
        t2.Src,
        t2.Note_1,
        t2.Note_2,
        'Rockauto' AS 'OE来源'
        FROM vehicle_table AS t1
        LEFT JOIN
        (SELECT *
        FROM rock_auto
        WHERE Type_Code={}) AS t2
        ON t1.Make = t2.Make
        AND t1.Model = t2.Model
        AND t1.`Year` = t2.`Year`
        AND t1.`Engine` = t2.`Engine`
        WHERE t2.Make IS NOT NULL
        OR t2.Model IS NOT NULL
        OR t2.`Year` IS NOT NULL
        OR t2.`Engine` IS NOT NULL;
'''.format(Type_Code)
cursor.execute(sql)
result = cursor.fetchall()
# des = cursor.description
# print([item[0] for item in des])

columns_ls = ['Model_US', 'Make', 'Year', 'Model', 'Engine', 'Brand', 'Part_Number',
              'OE_Numbers', 'Vehicle', 'Src', 'Note_1', 'Note_2', 'OE来源']
df = pd.DataFrame(result, columns=columns_ls)
cursor.close()
conn.close()

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 24
ws.column_dimensions['J'].width = 40
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 8
ws.freeze_panes = "A2"
font = Font(name='等线', size=9)
header_font = Font(name='等线', size=10, bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 20


wb.save('Data.xlsx')




