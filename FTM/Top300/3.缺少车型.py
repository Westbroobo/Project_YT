# -*- coding: utf-8 -*- 
# @Time ： 2024/7/13 11:24
# @Auth ： Westbroobo
# @File ：3.缺少车型.py


import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

conn = mysql.connector.connect(
    host="127.0.0.1",
    user="root",
    password="654321",
    database="data"
)
Type_Code = '5860'
cursor = conn.cursor()
sql = '''
        SELECT t1.*
        FROM vehicle_table_top300 AS t1
        LEFT JOIN
        (SELECT DISTINCT Make, Model, `Year`, `Engine`
        FROM rock_auto
        WHERE Type_Code={}) AS t2
        ON t1.Make = t2.Make
        AND t1.Model = t2.Model
        AND t1.`Year` = t2.`Year`
        AND t1.`Engine` = t2.`Engine`
        WHERE t2.Make IS NULL
        OR t2.Model IS NULL
        OR t2.`Year` IS NULL
        OR t2.`Engine` IS NULL;
'''.format(Type_Code)
cursor.execute(sql)
result = cursor.fetchall()
columns = ['Make', 'Model', 'Year', 'Engine']
df = pd.DataFrame(result, columns=columns)
cursor.close()
conn.close()

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 28

ws.freeze_panes = "A2"
font = Font(name='等线', size=9)
header_font = Font(name='等线', size=10, bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 20


wb.save('Lack_Vehicle.xlsx')


