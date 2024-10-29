# -*- coding: utf-8 -*- 
# @Time ： 2024/7/12 15:19
# @Auth ： Westbroobo
# @File ：1.车型逆向.py

from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sympy import rsolve

url = "mysql+pymysql://root:654321@127.0.0.1:3306?charset=utf8"
engine = create_engine(url=url, echo=True, future=True)
sql = '''
        SELECT CONCAT(Make,' ',Model) as Model_US, Year
        FROM data.vehicle_table
'''
with engine.connect() as conn:
    result = conn.execute(text(sql)).all()
    des = conn.execute(text(sql)).cursor.description
    print([item[0] for item in des])
    # df = pd.DataFrame(result, columns=['Model_US', 'Year'])
    # print(result)

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"
# for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#     ws.append(row)
# # ws.column_dimensions['A'].width = 20
# # ws.column_dimensions['B'].width = 15
#
# ws.freeze_panes = "A2"
# font = Font(name='等线', size=9)
# header_font = Font(name='等线', size=10, bold=True)
# fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
# alignment = Alignment(horizontal="center", vertical="center")
# border = Border(
#     left=Side(style='thin'),
#     right=Side(style='thin'),
#     top=Side(style='thin'),
#     bottom=Side(style='thin')
# )
# for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
#     for cell in row:
#         cell.font = font
#         cell.alignment = alignment
#         cell.border = border
# for cell in ws[1]:
#     cell.font = header_font
#     cell.fill = fill
#     cell.alignment = alignment
#     cell.border = border
# for row in range(1, ws.max_row + 1):
#     ws.row_dimensions[row].height = 20
#
#
# wb.save('Data.xlsx')


