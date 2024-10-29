"""
Filename:2.Amazon.py
Author: Westbroobo
Date: 2024/8/28
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import time
import pandas as pd
from tool import UA, Proxy
import requests
from lxml import etree
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'ASIN', 'Pic', 'Title', 'Grade', 'Ratings', 'Price', 'About this item', 'Json_Technical_Details', 'Url'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 9
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 25
ws.column_dimensions['I'].width = 25
ws.column_dimensions['J'].width = 45
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

wb.save('Amazon.xlsx')


df_menu = pd.read_excel('./menu.xlsx', header=0)
ASIN = df_menu['ASIN'].to_list()


def fetch_data(asin):
    num = 20
    for test in range(num):
        try:
            url = 'https://www.amazon.com/dp/' + asin + '?language=en_US'
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-language': 'zh-CN,zh;q=0.9', 'cache-control': 'max-age=0',
                'cookie': 'session-id=133-9846465-1707349; session-id-time=2082787201l; i18n-prefs=USD; ubid-main=131-3295320-4720951; lc-main=en_US; session-token=dwJreNOe5JcMf/IbYsol3uGwTF0waZDK5iNG60JAmkfvsfdWahMo+LUU0TgnAHCPG1mUFaQHhb/P328qEQ+bAvDrb4UD7aETzHqxU/98n2ZXr2Ybs+VMgnFQs3zMlOiZ/iUoVs4c48ujCmuJP4/VXET22VEwQEhT4E4ZcDPSlzrS+slyq5pRtZw0QykeiEm6OMvGyMiv8Oy7k2vOLN2roouojjp8op+2uht4n8LzRJA1qLLu+aPh9wFG5FZ79u7O0ADhJNKCPpXtmda1a8IaKNdTZmK/arJec3ZFpKbUQO9fDm8BZgMdS9K0AitsTga2rCXtqxluVQoJ86sO/BT3HCMofPi9+qOP; csm-hit=adb:adblk_no&t:1724723499898&tb:Z11K171KKNRR06EZFHC8+s-Z11K171KKNRR06EZFHC8|1724723499898',
                'device-memory': '8', 'downlink': '10', 'dpr': '1', 'ect': '4g', 'priority': 'u=0, i', 'rtt': '150',
                'sec-ch-device-memory': '8', 'sec-ch-dpr': '1',
                'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"', 'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"', 'sec-ch-ua-platform-version': '"10.0.0"',
                'sec-ch-viewport-width': '1015', 'sec-fetch-dest': 'document', 'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'none', 'sec-fetch-user': '?1', 'upgrade-insecure-requests': '1',
                'user-agent': UA.get_UA(), 'viewport-width': '1015', 'referer': 'https://www.amazon.com/'}
            rsp = requests.get(url=url, headers=headers, proxies=Proxy.get_Proxy_Requests())
            page_text = rsp.text
            status_code = rsp.status_code
            tree = etree.HTML(page_text)
            title = tree.xpath('//span[@id="productTitle"]/text()')[0].strip()
            img_Tag_Wrapper = tree.xpath('//*[@id="imgTagWrapperId"]')[0]
            img = img_Tag_Wrapper.xpath('./img')[0].get('src')
            try:
                grade = tree.xpath('//*[@id="acrPopover"]/span[1]/a/span/text()')[0].strip()
            except:
                grade = '-'
            try:
                ratings = tree.xpath('//*[@id="acrCustomerReviewText"]/text()')[0].strip()
            except:
                ratings = '-'
            try:
                try:
                    PriceAccordionT2 = tree.xpath('//*[@class="a-price a-text-normal aok-align-center reinventPriceAccordionT2"]')[0]
                    price = PriceAccordionT2.xpath('./span/text()')[0].strip()
                except:
                    PriceAccordionT2 = tree.xpath('//*[@class="a-section a-spacing-none aok-align-center aok-relative"]')[0]
                    price = PriceAccordionT2.xpath('./span/text()')[0].strip()
            except:
                price = '-'
            lis_item = tree.xpath('.//*[@class="a-unordered-list a-vertical a-spacing-mini"]/li')
            contents = []
            for li in lis_item:
                text = li.xpath('./span/text()')[0].strip()
                contents.append(text)
            content = '\n'.join(contents)
            Technical_Details = tree.xpath('//*[@id="productDetails_techSpec_section_1"]/tr')
            list_ = []
            for tr in Technical_Details:
                dict_ = {}
                key = tr.xpath('./th/text()')[0].strip().replace('\u200e', '')
                value = tr.xpath('./td/text()')[0].strip().replace('\u200e', '')
                dict_[key] = value
                list_.append(dict_)
            formatted_data = {str(index): item for index, item in enumerate(list_)}
            json_data = json.dumps(formatted_data)

            if status_code == 200 and len(list_) != 0:
            # if status_code == 200:
                wb = load_workbook(filename='./Amazon.xlsx')
                ws = wb['Data']
                serial = ws.max_row
                ws.append([serial, asin, img, title, grade, ratings, price, content, json_data, url])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 15
                for col in range(1, 11):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./Amazon.xlsx')
                wb.close()
                rsp.close()
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(asin, test+1, serial, len(ASIN)))
                break
        except:
            time.sleep(0.3)
            if test == num-1:
                print('{}【尝试次数：{}】-error'.format(asin, test))
            continue


pool = pool.Pool(10)
for asin in ASIN:
    pool.spawn(fetch_data, asin)
pool.join()

