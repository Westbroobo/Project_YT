"""
Filename:4.SMP_merge.py
Author: Westbroobo
Date: 2024/9/27
"""


import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def year_(year):
    if year in ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09'] or year in [str(i) for i in range(30)]:
        return '20' + year
    else:
        return '19' + year

def vehicle_handle_1(y):
    if '-' not in y.split(' ')[-1]:
        return year_(y)
    else:
        years = y.strip().split('-')
        year_1 = years[-1]
        year_1 = year_(year_1)
        year_2 = years[0]
        year_2 = year_(year_2)
        year_result = year_1 + '-' + year_2
        return year_result

def vehicle_handle_2(vehicle):
    try:
        list_ = vehicle.split('\n')
        content = []
        for l in list_:
            if ',' not in l:
                make_model = ' '.join(l.split(' ')[:-1])
                content.append(make_model+' '+vehicle_handle_1(l.split(' ')[-1]))
            else:
                l_ = l.split(', ')
                make_model = ' '.join(l_[0].split(' ')[:-1])
                content.append(make_model+' '+vehicle_handle_1(l_[0].split(' ')[-1]))
                for j in l_[1:]:
                    content.append(make_model + ' ' + vehicle_handle_1(j))
        return '\n'.join(content)
    except Exception as e:
        return ''

df_menu = pd.read_excel('./SMP_menu.xlsx')
df_detail = pd.read_excel('./SMP_detail.xlsx')
df = pd.merge(df_menu, df_detail, on='Part Number', how='left')
df.rename(columns={'Img': 'Pic', 'Buyer Guide': 'Vehicle'}, inplace=True)
df['No.'] = [i+1 for i in range(len(df))]
df = df[['No.', 'Part Number', 'Vehicle', 'Pic', 'Part Specifications']]
df['Vehicle'] = df['Vehicle'].apply(lambda x: vehicle_handle_2(x.replace('(', '').replace(')', '')))

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 60
ws.freeze_panes = "A2"
font = Font(name='微软雅黑', size=9)
header_font = Font(name='微软雅黑', size=9, color="FFFFFF", bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
ws.row_dimensions[1].height = 20
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 16.8

wb.save('./SMP.xlsx')
wb.close()



