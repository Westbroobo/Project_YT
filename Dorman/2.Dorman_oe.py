"""
Filename:2.Dorman_oe.py
Author: Westbroobo
Date: 2024/8/27
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import requests
import json
import time
from tool import UA, Proxy
from lxml import etree
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'Part_Number', 'Title', 'OE', 'Json_ProductSpec', 'Url', 'Img', 'Url_vehicle'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 40
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
wb.save('Dorman.xlsx')

df_menu = pd.read_excel('./Dorman_menu.xlsx', header=0, dtype=str).fillna('')
Url = df_menu['Url'].to_list()
Part_Number = df_menu['Part_Number'].to_list()
Title = df_menu['Title'].to_list()
Img = df_menu['Img'].to_list()
cookies = {
    '__uzma': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    '__uzmb': '1724736873',
    '__uzme': '1581',
    'SkinID': '1',
    '_ga': 'GA1.1.804590360.1724736878',
    '__ssds': '2',
    '__ssuzjsr2': 'a9be0cd8e',
    '__uzmbj2': '1724736879',
    '__uzmlj2': 'dgK7cWhhUpFotKaBe9b+DIMy2HmmHXkZf3CW3+GR5so=',
    '_gcl_au': '1.1.156773615.1724736881',
    'CultureLanguageDefaultSettingENG': 'AutoOff',
    '__uzmaj2': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    'Referrer': 'https%3A%2F%2Fvalidate.perfdrive.com%2F',
    'RecentActivity': '6837e550-6352-446b-8444-7558c34aa9ef',
    '_fbp': 'fb.1.1724739093350.535113685250122218',
    'ASP.NET_SessionId': '24nbdarsr0nlxkxcyi1vzqie',
    'gtm_session_start': '1724830265989',
    '_uetsid': '0edf79d0643611ef9bd401cd14792de9|1cs1tjb|2|fop|0|1700',
    '_ga_9Z678K7TDD': 'GS1.1.1724827848.5.1.1724830266.54.0.0',
    '_uetvid': '0edfa330643611efa3a8e3c008d5f315|aofj3i|1724830266762|5|1|bat.bing.com/p/insights/c/s',
    '__uzmdj2': '1724830267',
    '__uzmcj2': '626989415685',
    '__uzmfj2': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f8172473687960493387958-dd62d5ff7d9f7efc94',
    'uzmxj': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac52-172473687960493387958-50bc7e8ca8b6011794',
    '__uzmc': '5792713042201',
    '__uzmd': '1724830272',
    '__uzmf': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f8172473687306593399501-eab2e111b911d8d2130',
    'uzmx': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac52-172473687306593399501-b90bc5a67e2f6292130',
}
params = {
    'parttype': 'Radiator%20Fan%20Assembly',
    'origin': 'keyword',
}

def fetch_data(url):
    num = 100
    for test in range(num):
        part_number = Part_Number[Url.index(url)]
        title = Title[Url.index(url)]
        img = Img[Url.index(url)]
        try:
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-language': 'zh-CN,zh;q=0.9', 'cache-control': 'max-age=0', 'priority': 'u=0, i',
                'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"', 'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"', 'sec-fetch-dest': 'document', 'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'none', 'sec-fetch-user': '?1', 'upgrade-insecure-requests': '1',
                'user-agent': UA.get_UA(), 'referer': 'https://www.dormanproducts.com'}
            # params = {
            #     'parttype': re.compile(r'parttype=(.+?)&').findall(url)[0],
            #     'origin': re.compile(r'origin=(.*)').findall(url)[0],
            # }
            rsp = requests.get(url=url.split('?')[0], cookies=cookies, headers=headers,
                               params=params, proxies=Proxy.get_Proxy_Requests())
            status_code = rsp.status_code
            tree = etree.HTML(rsp.text)
            hidfDetailApp = tree.xpath('//div[@id="divDetailApps"]/input')[0].get('value')
            vehicle_url = 'https://www.dormanproducts.com/' + hidfDetailApp

            productOE = tree.xpath('//*[@id="productOE"]/div/table')
            list_oe = []
            for table in productOE:
                oe_trs = table.xpath('./tbody/tr')
                count = len(oe_trs)
                for c in range(count):
                    oe = table.xpath('./tbody/tr[{}]/th/text()'.format(int(c)+1))[0]
                    list_oe.append(oe)
            oe = ';'.join(list_oe)

            productSpec = tree.xpath('//*[@id="productSpec"]/div/table/tr')
            list_ = []
            for tr in productSpec:
                dict_ = {}
                key = tr.xpath('./th/text()')[0].strip()
                value = tr.xpath('./td/text()')[0].strip()
                dict_[key] = value
                list_.append(dict_)
            formatted_data = {str(index): item for index, item in enumerate(list_)}
            json_data = json.dumps(formatted_data)

            if status_code == 200:
                wb = load_workbook(filename='./Dorman.xlsx')
                ws = wb['Data']
                serial = ws.max_row
                ws.append([serial, part_number, title, oe, json_data, url, img, vehicle_url])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 15
                for col in range(1, 9):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./Dorman.xlsx')
                wb.close()
                rsp.close()
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(part_number, test + 1, serial, len(Url)))
                break
        except Exception as e:
            # print(e)
            time.sleep(0.3)
            if test == num - 1:
                print('{}【尝试次数：{}】-error'.format(part_number, test+1))
            continue

pool = pool.Pool(5)
for url in Url:
    pool.spawn(fetch_data, url)
pool.join()


