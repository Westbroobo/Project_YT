"""
Filename:3.Dorman_vehicle.py
Author: Westbroobo
Date: 2024/8/27
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import pandas as pd
import requests
from lxml import etree
from tool import UA, Proxy
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'Part_Number', 'Vehicle', 'Engine', 'Vehicle_Engine'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 65
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
wb.save('Dorman_vehicle.xlsx')

def list_cleand(lst):
    new_list = []
    lst.sort()
    for x, y in zip(lst, lst[1:]):
        if int(y) - int(x) > 1:
            new_list.append(lst[:lst.index(y)])
            lst = lst[lst.index(y):]
    new_list.append(lst)
    return new_list


def fetch_vehicle(detail_app_row, make_model_year_dict, year_make_model_configuration_ls, configuration_ls):
    for tr in detail_app_row:
        year = tr.xpath('./td[1]/text()')[0].strip()
        make = tr.xpath('./td[2]/text()')[0].strip()
        model = tr.xpath('./td[3]/text()')[0].strip()
        configuration = tr.xpath('./td[4]/text()')[0].strip()
        # position = tr.xpath('./td[5]/text()')[0].strip()
        # note = tr.xpath('./td[6]/text()')[0].strip()
        year_make_model_configuration = year + ' ' + make + ' ' + model + ' ' + configuration
        if year_make_model_configuration not in year_make_model_configuration_ls:
            year_make_model_configuration_ls.append(year_make_model_configuration)
        configuration_ls.append(configuration)
        if make not in make_model_year_dict:
            make_model_year_dict[make] = {model: [year]}
        else:
            if model not in make_model_year_dict[make]:
                make_model_year_dict[make][model] = [year]
            else:
                make_model_year_dict[make][model].append(year)
        make_model_year_dict[make][model] = list(set(make_model_year_dict[make][model]))
    return make_model_year_dict, year_make_model_configuration_ls, configuration_ls


def dict_to_str(make_model_year_dict):
    make_model_year_ls = []
    for make in make_model_year_dict:
        for model in make_model_year_dict[make]:
            if len(make_model_year_dict[make][model]) == 1:
                content = make + ' ' + model + ' ' + str(make_model_year_dict[make][model][0])
                make_model_year_ls.append(content)
            elif int(max(make_model_year_dict[make][model])) - int(min(make_model_year_dict[make][model])) == len(
                    make_model_year_dict[make][model]) - 1:
                content = make + ' ' + model + ' ' + str(min(make_model_year_dict[make][model])) \
                          + '-' + str(max(make_model_year_dict[make][model]))
                make_model_year_ls.append(content)
            else:
                new_list = list_cleand(make_model_year_dict[make][model])
                for n in new_list:
                    if len(n) == 1:
                        content = make + ' ' + model + ' ' + str(n[0])
                        make_model_year_ls.append(content)
                    else:
                        content = make + ' ' + model + ' ' + str(min(n)) + '-' + str(max(n))
                        make_model_year_ls.append(content)
    make_model_year_ls = list(set(make_model_year_ls))
    make_model_year_ls.sort()
    make_model_year = '\n'.join(make_model_year_ls)
    return make_model_year


cookies = {
    '__uzma': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    '__uzmb': '1724736873',
    '__uzme': '1581',
    'SkinID': '1',
    '_ga': 'GA1.1.804590360.1724736878',
    '__ssds': '2',
    '__ssuzjsr2': 'a9be0cd8e',
    '__uzmbj2': '1724736879',
    '__uzmlj2': 'dgK7cWhhUpFotKaBe9b+DIMy2HmmHXkZf3CW3+GR5so=',
    '_gcl_au': '1.1.156773615.1724736881',
    'CultureLanguageDefaultSettingENG': 'AutoOff',
    '__uzmaj2': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    'Referrer': 'https%3A%2F%2Fvalidate.perfdrive.com%2F',
    'RecentActivity': '6837e550-6352-446b-8444-7558c34aa9ef',
    '_fbp': 'fb.1.1724739093350.535113685250122218',
    'ASP.NET_SessionId': '24nbdarsr0nlxkxcyi1vzqie',
    'gtm_session_start': '1724830265989',
    '_uetsid': '0edf79d0643611ef9bd401cd14792de9|1cs1tjb|2|fop|0|1700',
    '_uetvid': '0edfa330643611efa3a8e3c008d5f315|aofj3i|1724830280090|6|1|bat.bing.com/p/insights/c/s',
    'gtm_session_threshold': 'true',
    '_ga_9Z678K7TDD': 'GS1.1.1724827848.5.1.1724830397.60.0.0',
    '__uzmc': '4144217877868',
    '__uzmd': '1724831323',
    '__uzmf': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f8172473687306594450915-d2f2b2de17a2b3e4178',
    'uzmx': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac52-172473687306594450915-aba5413a44427828178',
    '__uzmcj2': '2740913999847',
    '__uzmdj2': '1724831324',
    '__uzmfj2': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f8172473687960494444976-b4a395943920dbfe139',
    'uzmxj': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac52-172473687960494444976-f9407aa598262fc2139',
}
headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cache-control': 'max-age=0',
    'priority': 'u=0, i',
    'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': UA.get_UA(),
    'referer': 'https://www.dormanproducts.com'
}
df_menu = pd.read_excel('./Dorman.xlsx', header=0)
Part_Number = df_menu['Part_Number'].to_list()
Url_vehicle = df_menu['Url_vehicle'].to_list()

def fetch_data(url):
    num = 100
    for test in range(num):
        part_number = Part_Number[Url_vehicle.index(url)]
        try:
            session = requests.session()
            Captcha = []
            make_model_year_dict = {}
            year_make_model_configuration_ls = []
            configuration_ls = []
            while True:
                rsp = session.get(url=url, cookies=cookies, headers=headers, proxies=Proxy.get_Proxy_Requests())
                if 'ShieldSquare Captcha' in rsp.text:
                    print(Captcha[0])
                status_code = rsp.status_code
                tree = etree.HTML(rsp.text)
                detail_app_row = tree.xpath('//*[@class="detail-app-row"]')
                make_model_year_dict, year_make_model_configuration_ls, configuration_ls = (fetch_vehicle
                                                                                            (detail_app_row, make_model_year_dict, year_make_model_configuration_ls, configuration_ls))
                next_button = tree.xpath('//*[@id="pagingBottom_nextButton"]')
                if len(next_button) == 0:
                    session.close()
                    break
                else:
                    href = next_button[0].get('href')
                    url = 'https://www.dormanproducts.com' + href
            configuration_ls = list(set(configuration_ls))
            year_make_model_configuration_ = '\n'.join(year_make_model_configuration_ls).strip()
            configuration_ = '\n'.join(configuration_ls).strip()
            make_model_year = dict_to_str(make_model_year_dict)
            if status_code == 200:
                wb = load_workbook(filename='./Dorman_vehicle.xlsx')
                ws = wb['Data']
                serial = ws.max_row
                ws.append([serial, part_number, make_model_year, configuration_, year_make_model_configuration_])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 20
                for col in range(1, 6):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./Dorman_vehicle.xlsx')
                wb.close()
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(part_number, test + 1, serial, len(Url_vehicle)))
                break
        except Exception as e:
            # print(e)
            time.sleep(0.3)
            if test == num - 1:
                print('{}【尝试次数：{}】-error'.format(part_number, test))
            continue

pool = pool.Pool(5)
for url in Url_vehicle:
    pool.spawn(fetch_data, url)
pool.join()


