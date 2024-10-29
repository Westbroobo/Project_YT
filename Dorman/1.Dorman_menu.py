"""
Filename:1.Dorman_menu.py
Author: Westbroobo
Date: 2024/8/27
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import requests
import time
from tool import UA, Proxy
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

cookies = {
    '__uzma': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    '__uzmb': '1724736873',
    '__uzme': '1581',
    'ASP.NET_SessionId': 'eexfphuw1k00uctly5pr0sva',
    'SkinID': '1',
    '_ga': 'GA1.1.804590360.1724736878',
    'gtm_session_start': '1724736878873',
    '__ssds': '2',
    '__ssuzjsr2': 'a9be0cd8e',
    '__uzmbj2': '1724736879',
    '__uzmlj2': 'dgK7cWhhUpFotKaBe9b+DIMy2HmmHXkZf3CW3+GR5so=',
    '_gcl_au': '1.1.156773615.1724736881',
    'CultureLanguageDefaultSettingENG': 'AutoOff',
    '__uzmaj2': '495a586d-45b4-4fe3-9cd0-2e05065c5516',
    'gtm_session_threshold': 'true',
    'tempSearchCookie': 'Radiator+Fan+Assembly=https%3a%2f%2fwww.dormanproducts.com%2fgsearch.aspx%3ftype%3dkeyword%26origin%3dkeyword%26parttype%3dRadiator+Fan+Assembly%26start%3d100%26num%3d100',
    '__uzmcj2': '294113194664',
    '__uzmdj2': '1724737123',
    '__uzmfj2': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f81724736879604243402-8450059c734c59fa31',
    'uzmxj': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac51-1724736879604243402-caaea42462ba836c31',
    '_ga_9Z678K7TDD': 'GS1.1.1724736878.1.1.1724737123.14.0.0',
    '_uetsid': '0edf79d0643611ef9bd401cd14792de9|1cs1tjb|2|foo|0|1700',
    '_uetvid': '0edfa330643611efa3a8e3c008d5f315|1myi0sc|1724737123461|8|1|bat.bing.com/p/insights/c/e',
    '__uzmc': '355014393688',
    '__uzmd': '1724737189',
    '__uzmf': '7f6000ee00a447-ccad-4fa8-a3e4-a3f9ec28b1f81724736873065316464-0c99a7d1e4a5128643',
    'uzmx': '7f9000e68c77d0-480d-4cb5-adc4-7e3b19f99ac51-1724736873065316464-1f24e507261661a443',
}

wb = Workbook()
ws = wb.active
ws.title = "menu"
ws.append(['Part_Number', 'Title', 'Img', 'Url'])
ws.freeze_panes = "A2"
wb.save(filename='./Dorman_menu.xlsx')
wb.close()
def fetch_data(start, page):
    num = 100
    for test in range(num):
        try:
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-language': 'zh-CN,zh;q=0.9', 'cache-control': 'max-age=0', 'priority': 'u=0, i',
                'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"', 'sec-fetch-dest': 'document', 'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'none', 'sec-fetch-user': '?1', 'upgrade-insecure-requests': '1',
                'user-agent': UA.get_UA(),
                'referer': 'https://www.dormanproducts.com'}
            params = {
                'type': 'keyword',
                'origin': 'keyword',
                'parttype': 'Radiator%20Fan%20Assembly',
                'start': start,
                'num': '100',
            }
            rsp = requests.get('https://www.dormanproducts.com/gsearch.aspx',
                               params=params, cookies=cookies, headers=headers, proxies=Proxy.get_Proxy_Requests())
            status_code = rsp.status_code
            tree = etree.HTML(rsp.text)
            wb = load_workbook(filename='./Dorman_menu.xlsx')
            ws = wb['menu']
            searchItems = tree.xpath('//*[@class="searchItems"]')
            for item in searchItems:
                href = item.xpath('./div[1]/a')[0].get('href')
                detail_url = 'https://www.dormanproducts.com/' + href
                img = item.xpath('./div[1]/a/img')[0].get('src')
                part_number = item.xpath('./div[2]/h2/a/span/text()')[0]
                title = item.xpath('./div[2]/div[2]/h4/text()')[0]
                ws.append([part_number, title, img, detail_url])
            ws.column_dimensions['A'].width = 13
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 60
            ws.column_dimensions['D'].width = 80
            font = Font(name='等线', size=9)
            header_font = Font(name='等线', size=10, bold=True)
            fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = fill
                cell.alignment = alignment
                cell.border = border
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 15
            wb.save(filename='./Dorman_menu.xlsx')
            wb.close()
            rsp.close()
            if status_code == 200:
                print('第{}页【尝试次数：{}】-success'.
                      format(page, test + 1))
                break
        except:
            time.sleep(0.3)
            if test == num - 1:
                print('第{}页【尝试次数：{}】-error'.format(page, test))
            continue


pool = pool.Pool(3)
for i in range(3):
    page = int(i)+1
    start = str(i*100)
    pool.spawn(fetch_data, start, page)
pool.join()


