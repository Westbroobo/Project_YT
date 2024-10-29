"""
Filename:2.fetch_all.py
Author: Westbroobo
Date: 2024/8/29
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df_menu = pd.read_excel('./Json_Data.xlsx', header=0)
Json_Data = df_menu['Json_Data'].tolist()

df = pd.DataFrame()
for json_data in Json_Data:
    data_dict = json.loads(json_data)
    dict_ = {}
    for d in data_dict.values():
        dict_.update(d)
    df_ = pd.DataFrame([dict_])
    df = pd.concat([df, df_], ignore_index=True).fillna('')

wb = Workbook()
ws = wb.active
ws.title = "Result"
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15
ws.column_dimensions['M'].width = 15
ws.column_dimensions['N'].width = 15
ws.column_dimensions['O'].width = 15
ws.freeze_panes = "A2"
font = Font(name='等线', size=9)
header_font = Font(name='等线', size=10, bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 15

wb.save('Result.xlsx')

