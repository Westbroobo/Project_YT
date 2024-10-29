"""
Filename:2.Detail.py
Author: Westbroobo
Date: 2024/9/10
"""
import time
import requests
from lxml import etree
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['Url', 'Product Id', 'Title', 'Price', 'OE', 'OE-clean', 'Json_Product_Detail', 'Vehicle_Make'])
df_menu = pd.read_excel('./menu.xlsx')
href = df_menu['href'].tolist()
cookies = {
    'cf_clearance': '6EBthzmiWw0har7Ms9Q.9dhlaXelN5p_OKI4BDSAdRw-1725690818-1.2.1.1-EAAF737qbpLyhpqP.DyF9l4D7q9hIFQlSHN5fCRy3r6fS5XSu6z8MRqOpVGgfw9DmPOY0GLv1aqySY_T8M0vqFkJdIL6Wh9Ih8GQSkDBgPPdfhGN1nqLHByfImDHkb5oWzSj2VnmAImf602DVI9eyDz6GpbS761BsehTpSD.JVP6bysSGI9OZIa3QLiYf4EqamSe3jpQCbeH0tTSBgye8GYcR2JVQGi57GjCsCASH8z3jYdPl8kb_RytqBptNedts9PZyWCepq7PqcU2bdhjQwR1OYs65EeJZQjjrr19tF482IqyO6NWb9WqNqvqSNAH6Elm6k14FX6J0lo55z4IA2j0bs1JFefuv_NC7kjczSaED2tsfMZUDxrF5sbnQbzkONYFbsE2CuB0X6cCroB6Q4ky25vi1WXW5X8fjkcS380d416ZvatPtadCDcebpJEH',
    'kmtx_sync': '3893828442274565296',
    'kmtx_sync': '3893828442274565296',
    '_fbp': 'fb.1.1725690835852.43780976221083146',
    'INGRESSCOOKIE': '1725935102.938.14141.51617|f89e2d2acc6cd8c158ba7aaea6f7fb53',
    'clientCode': '243324',
    '_gcl_au': '1.1.308737845.1725935108',
    '_ga': 'GA1.1.1170940819.1725935105',
    'tb_visit_counter_direkt_de': 'eyJpdiI6Ild3TjdlYU5rUGtXWDdLclQxeG1kTnc9PSIsInZhbHVlIjoidFZzN05OQ0c0eERFeHFrNnNINVk2UVAwNDZvZmw5WnZ5VFhWWjlCOWIraTFQb3d2K0dKcmxHUGdqQkZ3WHprayIsIm1hYyI6IjNhMTZmMmJhNDRhNmNlYWQ2OGQ3YjMxMjZjOGQzZDA5ODg5MDFjMTI4MDdlYzlmMGVkZDhkNjQxMTA3ZDg5YmEiLCJ0YWciOiIifQ%3D%3D',
    '__cf_bm': 'lq2G3LcYWD5mhLwRiTAyxQ4hoxejdtRkO1ibaZi_3b8-1725947143-1.0.1.1-P3pN1vxx_s7XWmmUSWEUtYD0EfUPXHWBTyxT03hjmHTvCRpXQ5iTub8du47r0URcEhtTIcvI.kbd.1jNs63rcvfLnm0bKIlzHjF._ecCPnk',
    'kmtx_sync_pi': '99be6966-4d98-4036-8a78-9768bd240cdb',
    '_uetsid': 'e6fbd1c06f1b11ef871c91300f4033c2',
    '_uetvid': '26dcbdb06ce311ef8d2a756a0cf7b6c1',
    '_ga_DHWDNTWLEQ': 'GS1.1.1725947154.3.1.1725947538.50.0.1794647297',
    'XSRF-TOKEN': 'eyJpdiI6InRtRWdQWHNDNWdMak9FdDNpbHNOUGc9PSIsInZhbHVlIjoiQjZpaG1MQ2dLY1FZRHVOMEpORlBTbGZZaFM5RnpwMXV2aDMzZDRFUnVrNHZXTk1ENk5VOFBWYjREVE1LemxwZVdhek1zNGc5aThOYjZLRXNraUlSSk5jM2lCVnNtNFpEeCtRc0RlVTlVanMyeStSbEkvV0ZGbERiZ1NOZTRvR2YiLCJtYWMiOiI0MDA3N2VhNjI0NWZhNzdjZGFkMTIxNjJjZDhjNDI5MjQzYjYwYjBiOTI0NGY1NDZmYTZmYzFlNTQ0ZGZlNjU0IiwidGFnIjoiIn0%3D',
    'l': 'eyJpdiI6IldnWTFUWlNEVDJkRldrdTZoQ1NKeHc9PSIsInZhbHVlIjoiNVBob2ZaQTM4Z1gwUW5TYnljZ3V5MU1XYWVNTHhDajF5dFBIeWZPakZkZFZnZ21ZS3lPSm4wRnZCaGhMalducTZaU2V2Z2ZUQU9pdU1vYzVXdkdyakZGMmswbWhJRUh2WkN5eFRKRDYxSGNQelZTRG4xK0xRaytwZUVHMC8zTUciLCJtYWMiOiJmMzA1YzUzYmViMmE0NGQ0ZmRkNzNiYWMyNmNiNWMzNzliMjViMWI3ZTVlNTI3MTM3NTEzOTUzNWI1Y2ZlNjk5IiwidGFnIjoiIn0%3D',
}
headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cache-control': 'max-age=0',
    'priority': 'u=0, i',
    'sec-ch-ua': '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
    'sec-ch-ua-arch': '"x86"',
    'sec-ch-ua-bitness': '"64"',
    'sec-ch-ua-full-version': '"128.0.6613.120"',
    'sec-ch-ua-full-version-list': '"Chromium";v="128.0.6613.120", "Not;A=Brand";v="24.0.0.0", "Google Chrome";v="128.0.6613.120"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-model': '""',
    'sec-ch-ua-platform': '"Windows"',
    'sec-ch-ua-platform-version': '"10.0.0"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36',
}
error_url = []
for url in href:
    try:
        rsp = requests.get(url=url, cookies=cookies, headers=headers)
        tree = etree.HTML(rsp.text)
        productId = tree.xpath('//*[@name="articleId"]')[0].get('value')
        title = tree.xpath('//*[@class="product__title-link"]/text()')[0].strip()
        price = tree.xpath('//*[@class="product__new-price"]/text()')[0].strip()
        # oe 部分
        try:
            lis = tree.xpath('//*[@class="product-oem__list"]/ul/li')
            list_oe = []
            list_oe_cleaned = []
            for li in lis:
                try:
                    li_oe = li.xpath('./a/text()')[0].strip()
                except:
                    li_oe = li.xpath('./text()')[0].strip()
                li_oe_c = li_oe[li_oe.index('OE'):].replace('OE', '')
                if li_oe_c[0] == '-':
                    li_oe_c = li_oe_c[1:]
                list_oe.append(li_oe)
                list_oe_cleaned.append(li_oe_c)
            oes = '\n'.join(list_oe)
            oes_cleaned = '\n'.join(list_oe_cleaned)
        except:
            oes_cleaned = ''
            oes = ''
        # 参数部分
        product_table_row = tree.xpath('//div[@class="product__desc-content"]/div[1]/table/tr')
        list_ = []
        for row in product_table_row:
            dict_ = {}
            key = row.xpath('./td[1]/text()')[0].strip()
            value = row.xpath('./td[2]/text()')[0].strip()
            dict_[key] = value
            list_.append(dict_)
        formatted_data = {str(index): item for index, item in enumerate(list_)}
        json_data = json.dumps(formatted_data)
        # 车型部分
        compatibility_maker_title = tree.xpath('//*[@class="compatibility__maker-title"]')
        dict__ = {}
        for maker in compatibility_maker_title:
            make = maker.xpath('./text()')[0].strip()
            makerId = maker.get('data-maker-id')
            dict__[make] = makerId
        df_transit = pd.DataFrame([{'Url': url, 'Product Id': productId, 'Title': title,
                                    'Price': price, 'OE': oes, 'OE-clean': oes_cleaned,
                                    'Json_Product_Detail': json_data, 'Vehicle_Make': str(dict__)}])
        df = pd.concat([df, df_transit], ignore_index=True).fillna('')
        rsp.close()
        print('success-{}, error-{}, processing-{}'.format(len(df), len(error_url), len(href)-len(df)-len(error_url)))
    except:
        error_url.append(error_url)
        time.sleep(1)
        print('success-{}, error-{}, processing-{}'.format(len(df), len(error_url), len(href)-len(df)-len(error_url)))
        continue

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 28
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 28
ws.freeze_panes = "A2"
font = Font(name='微软雅黑', size=9)
header_font = Font(name='微软雅黑', size=9, color="FFFFFF", bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
ws.row_dimensions[1].height = 20
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 16.8

wb.save('./detail.xlsx')
wb.close()
