"""
Filename:4.result.py
Author: Westbroobo
Date: 2024/9/27
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df_input = pd.read_excel('./Vehicle.xlsx')
df_detail =  pd.read_excel('./detail.xlsx')
df_ = pd.merge(df_detail.iloc[:,:7], df_input, on='Product Id', how='left').fillna('-')

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df_, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 55
ws.freeze_panes = "A2"
font = Font(name='微软雅黑', size=9)
header_font = Font(name='微软雅黑', size=9, color="FFFFFF", bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
ws.row_dimensions[1].height = 20
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 16.8

wb.save('./result.xlsx')
wb.close()

