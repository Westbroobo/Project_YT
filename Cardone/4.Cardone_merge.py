"""
Filename:4.Cardone_merge.py
Author: Westbroobo
Date: 2024/9/27
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
df_menu = pd.read_excel('./Cardone_menu.xlsx')
df_oe = pd.read_excel('./Cardone_oe.xlsx')
df_technical = pd.read_excel('./Cardone_Technical.xlsx')

def vehicle_handle_1(vehicle):
    if '-' not in vehicle.split(' ')[-1]:
        return vehicle
    else:
        make_model = ' '.join(vehicle.strip().split(' ')[:-1])
        years = vehicle.strip().split(' ')[-1]
        year_1 = years[years.index('-')+1:]
        if year_1 in ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09'] or year_1 in [str(i) for i in range(30)]:
            year_1 = '20' + year_1
        else:
            year_1 = '19' + year_1
        year_2 = years[:years.index('-')]
        make_model_year = make_model + ' ' + year_1 + '-' + year_2
        return make_model_year


def vehicle_handle_2(vehicle):
    try:
        list_ = vehicle.split('\n')
        content = []
        for l in list_:
            if ',' not in l:
                content.append(vehicle_handle_1(l))
            else:
                l_ = l.split(', ')
                make = l_[0].split(' ')[0]
                if make in ['Land', 'Alfa', 'Aston']:
                    make = ' '.join(l_[0].split(' ')[:2])
                content.append(vehicle_handle_1(l_[0]))
                for j in l_[1:]:
                    content.append(make + ' ' + vehicle_handle_1(j))
        return '\n'.join(content)
    except:
        return ''

df = pd.merge(df_menu, df_oe, on='Part_Number', how='left')
df_ = pd.merge(df, df_technical, on='Part_Number', how='left')
df_['No.'] = [i+1 for i in range(len(df_))]
df_ = df_[['No.', 'Part_Number', 'Vehicle', 'OE', 'Technical', 'img_url', 'product_url']]
df_.rename(columns={'img_url': 'Img', 'product_url': 'Url'}, inplace=True)
df_['Vehicle'] = df_['Vehicle'].apply(lambda x: vehicle_handle_2(x))

wb = Workbook()
ws = wb.active
ws.title = "Data"
for r_idx, row in enumerate(dataframe_to_rows(df_, index=False, header=True), 1):
    ws.append(row)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 45
ws.column_dimensions['G'].width = 45
ws.freeze_panes = "A2"
font = Font(name='微软雅黑', size=9)
header_font = Font(name='微软雅黑', size=9, color="FFFFFF", bold=True)
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
ws.row_dimensions[1].height = 20
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 16.8

wb.save('./Cardone.xlsx')
wb.close()

