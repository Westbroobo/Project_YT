"""
Filename:2.Detail.py
Author: Westbroobo
Date: 2024/11/6
"""
from gevent import monkey, pool
monkey.patch_all(thread=False)
import requests, re, time
from tool import UA, Proxy
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'PageNumber', 'CompanyName', 'Pstands', 'ExhibitorUrlRewrite',
                           'Country', 'City', 'Street', 'Zip', 'Tel', 'Description', 'AdLabel', 'Url'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='微软雅黑', size=9, bold=True)
font = Font(name='微软雅黑', size=9)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 9
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 18
ws.column_dimensions['K'].width = 25
ws.column_dimensions['L'].width = 25
ws.column_dimensions['M'].width = 25
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

wb.save('Detail.xlsx')
df_menu = pd.read_excel('./menu.xlsx')
ExhibitorUrlRewrite = df_menu['ExhibitorUrlRewrite'].to_list()
CompanyName = df_menu['CompanyName'].to_list()
Pstands = df_menu['Pstands'].to_list()
PageNumber = df_menu['PageNumber'].to_list()

def fetch_data(exhibitor):
    num = 10
    for test in range(num):
        try:
            headers = {
                'Accept': 'application/json, text/plain, */*',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Connection': 'keep-alive',
                'Origin': 'https://automechanika-shanghai.hk.messefrankfurt.com',
                'Referer': 'https://automechanika-shanghai.hk.messefrankfurt.com/',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'same-site',
                'User-Agent': UA.get_UA(),
                'apikey': 'LXnMWcYQhipLAS7rImEzmZ3CkrU033FMha9cwVSngG4vbufTsAOCQQ==',
                'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
            }
            url = (f'https://api.messefrankfurt.com/service/esb_api/exhibitor-service/api/2.1/'
                   f'public/exhibitor/profile/zh-CN/{exhibitor}/AUTOMECHANIKASHANGHAI')
            rsp = requests.get(url=url, headers=headers, proxies=Proxy.get_Proxy_Requests())
            status_code = rsp.status_code
            try:
                city = re.findall('\"city\":\"(.+?)\",', rsp.text)[0]
            except:
                city = ''
            try:
                street = re.findall('\"street\":\"(.+?)\",', rsp.text)[0]
            except:
                street = ''
            try:
                zip = re.findall('\"zip\":\"(.+?)\",', rsp.text)[0]
            except:
                zip = ''
            try:
                tel = re.findall('\"tel\":\"(.+?)\",', rsp.text)[0]
            except:
                tel = ''
            try:
                label = re.findall('\"label\":\"(.+?)\"},', rsp.text)[0]
            except:
                label = ''
            try:
                description = re.findall('\"description\":(.+?),\"specification', rsp.text)[0]
                description_ = re.findall('\"text\":\"(.+?)\"}', description)[0]
                description_ = (description_.replace('<P>', '').replace('</P>', '\n').replace('<SPAN>', '').
                                replace('</SPAN>', '').replace('<UL>','').replace('</UL>','').replace('<LI>','').
                                replace('</LI>','\n').replace('<BR />', '').replace('\\t', '').replace('<OL>', '').
                                replace('</OL>', '').replace('&ldquo;', '"').replace('&rdquo;', '"').
                                replace('<STRONG>', '').replace('</STRONG>', '').replace('&nbsp;', '').
                                replace('<strong>', '').replace('</strong>', '').replace('&quot;', '').strip())
            except:
                description_ =''
            try:
                subCategories = re.findall('\"subCategories\":\\[(.+?)],', rsp.text)[0]
                adLabel = re.findall('\"adLabel\":\"(.+?)\",', subCategories)
                adLabel = '\n'.join(adLabel)
            except:
                adLabel = ''
            if status_code == 200:
                wb = load_workbook(filename='./Detail.xlsx')
                ws = wb['Data']
                page_number = PageNumber[ExhibitorUrlRewrite.index(exhibitor)]
                company_name = CompanyName[ExhibitorUrlRewrite.index(exhibitor)]
                pstands = Pstands[ExhibitorUrlRewrite.index(exhibitor)]
                serial = ws.max_row
                ws.append([serial, page_number, company_name, pstands, exhibitor, label,
                           city, street, zip, tel, description_, adLabel, url])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 16.8
                for col in range(1, 14):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./Detail.xlsx')
                wb.close()
                rsp.close()
                print(f'{exhibitor}【尝试次数：{test+1}】-success，进度：{serial}/{len(ExhibitorUrlRewrite)}')
                break
        except Exception as e:
            # print(e)
            time.sleep(0.3)
            if test == num - 1:
                print(f'{exhibitor}【尝试次数：{test}】-error')
            continue
pool = pool.Pool(15)
for exhibitor in ExhibitorUrlRewrite:
    pool.spawn(fetch_data, exhibitor)
pool.join()


