"""
Filename:3-1.fetch_vehicle1.py
Author: Westbroobo
Date: 2024/9/6
"""

from gevent import monkey, pool
monkey.patch_all(thread=False)
import requests
from tool import UA, Proxy
from lxml import etree
import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.DataFrame(columns=['No.', 'Search_Number', 'Category', 'Brand', 'Part_Number', 'OE', 'Note', 'Img', 'Url'])
wb = Workbook()
ws = wb.active
ws.title = "Data"
header_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 40
ws.column_dimensions['I'].width = 40
ws.freeze_panes = "A2"
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

wb.save('search_result.xlsx')

def fetch_data(p):
    num = 100
    for test in range(num):
        try:
            partnum_search = str(p)
            url = 'https://www.rockauto.com/en/partsearch/?partnum={}'.format(partnum_search)
            rsp = requests.get(url=url, headers=UA.get_User_Agent_Requests(), proxies=Proxy.get_Proxy_Requests())
            status_code = rsp.status_code
            tree = etree.HTML(rsp.text)
            nnormal = tree.xpath('./descendant::td[@class="nlabel"]/a/text()')[0]
            listing_container = tree.xpath('//div[@class="listing-container-border"]/div/table/tbody')
            if len(listing_container) == 0:
                wb = load_workbook(filename='./search_result.xlsx')
                ws = wb['Data']
                serial = ws.max_row
                ws.append([serial, partnum_search, '-', '-', '-', '-', '-', '-', url])
                last_row = ws.max_row
                ws.row_dimensions[last_row].height = 15
                for col in range(1, 10):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                wb.save(filename='./search_result.xlsx')
                wb.close()
            else:
                for tbody in listing_container[1:]:
                    brand = tbody.xpath('./descendant::span[@class="listing-final-manufacturer "]/text()')[0]
                    part_number = tbody.xpath('./descendant::div[@class="listing-text-row-moreinfo-truck"]/span[@title="Buyer\'s Guide"]/text()')[0]
                    category = tbody.xpath('./descendant::div[@class="listing-text-row"]/span/span/text()')[0].replace('Category:', '').strip()
                    oes = tbody.xpath('./descendant::span[@title="Replaces these Alternate/ OE Part Numbers"]/text()')
                    oe = ';'.join(oes).replace(', ', ';')
                    notes = tbody.xpath('./descendant::div[@class="listing-text-row-moreinfo-truck"]'
                                         '/span[@class="span-link-underline-remover"]/text()')
                    note = ';'.join(notes).replace(', ', ';')
                    try:
                        src = 'https://www.rockauto.com' + \
                              tbody.xpath('./descendant::img[@class=" listing-inline-image"]/@src')[0]
                    except IndexError:
                        src = ''
                    wb = load_workbook(filename='./search_result.xlsx')
                    ws = wb['Data']
                    serial = ws.max_row
                    ws.append([serial, partnum_search, category, brand, part_number, oe, note, src, url])
                    last_row = ws.max_row
                    ws.row_dimensions[last_row].height = 15
                    for col in range(1, 10):
                        cell = ws.cell(row=last_row, column=col)
                        cell.font = font
                        cell.alignment = alignment
                        cell.border = border
                    wb.save(filename='./search_result.xlsx')
                    wb.close()
            if status_code == 200:
                rsp.close()
                df_temp = pd.read_excel('search_result.xlsx')
                ls_partnum_search = list(set(df_temp['Search_Number'].tolist()))
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(p, test + 1, len(ls_partnum_search), len(partnums)))
                break
        except Exception as e:
            # print(e)
            time.sleep(0.3)
            if test == num-1:
                print('{}【尝试次数：{}】-error'.format(p, test))
            continue

df = pd.read_excel('./search_number.xlsx')
partnums = df['Search Number'].tolist()
pool = pool.Pool(10)
for p in partnums:
    pool.spawn(fetch_data, p)
pool.join()


