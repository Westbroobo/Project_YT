# -*- coding: utf-8 -*- 
# @Time ： 2024/7/4 15:05
# @Auth ： Westbroobo
# @File ：2-1.Rockauto_basic.py

from gevent import monkey, pool
monkey.patch_all(thread=False)
import time
import requests
from lxml import etree
from tool import UA, Proxy
import re
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
start = time.perf_counter()

df_menu = pd.read_excel('./Rockauto_menu.xlsx', header=0, dtype=str).fillna('')
Url = df_menu['Url'].to_list()
Part_Number = df_menu['Part_Number'].to_list()
wb = xlsxwriter.Workbook('./Rockauto_basic.xlsx')
ws = wb.add_worksheet('Sheet1')

ws.set_column('A:A', 9)
ws.set_column('B:B', 15)
ws.set_column('C:C', 18)
ws.set_column('D:D', 25)
ws.set_column('E:E', 45)
ws.set_column('F:F', 20)
ws.set_column('G:G', 18)
ws.set_column('H:H', 55)
headings = ['No.', 'Brand', 'Part_Number', 'payload', 'Pic', 'OE', 'Note', 'Url']
head_format = wb.add_format({
    'bold': 1,
    'fg_color': 'cyan',
    'align': 'center',
    'font_name': '等线',
    'size': 10,
    'valign': 'vcenter',
    'border': 1
})
ws.write_row('A1', headings, head_format)
ws.freeze_panes(1, 0)
ws.set_row(0, 20)
wb.close()
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))


def fetch_data(url):
    num = 50
    for test in range(num):
        try:
            rsp_detail = requests.get(url=url, headers=UA.get_User_Agent_Requests(), proxies=Proxy.get_Proxy_Requests())
            status_code_1 = rsp_detail.status_code
            tree_1 = etree.HTML(rsp_detail.text)
            brand = tree_1.xpath('./descendant::span[@class="listing-final-manufacturer "]/text()')[0]
            part_number = tree_1.xpath('./descendant::div[@class="listing-text-row-moreinfo-truck"]/'
                                       'span[@title="Buyer\'s Guide"]/text()')[0]
            oes = tree_1.xpath('./descendant::span[@title="Replaces these Alternate/ OE Part Numbers"]/text()')
            oe = ';'.join(oes).replace(', ', ';')
            notes = tree_1.xpath('./descendant::div[@class="listing-text-row-moreinfo-truck"]'
                                 '/span[@class="span-link-underline-remover"]/text()')
            note = ';'.join(notes).replace(', ', ';')
            try:
                src = 'https://www.rockauto.com' + \
                      tree_1.xpath('./descendant::img[@class=" listing-inline-image"]/@src')[0]
            except IndexError:
                src = ''
            href = tree_1.xpath('./descendant::a[@class="ra-btn ra-btn-moreinfo"]/@href')[0]
            part_key = '"' + re.search(r'pk=(\d+)', href).group(1) + '"'
            part_type = '"' + url.split(',')[-1] + '"'
            wh_part_num = '"' + brand + ' ' + part_number + '"'
            opts = '{"0-0-0-1": {' + '''
                      "warehouse": "72901",
                      "whpartnum": {},
                      "optionlist": "0",
                      "paramcode": "0",
                      "notekey": "0",
                      "multiple": "1"
                      '''.format(wh_part_num) + '}}'
            listing_data_essential = '{' + '''
                  "groupindex": "0",
                  "carcode": 0,
                  "parttype": {},
                  "partkey": {},
                  "opts": {}
                  '''.format(part_type, part_key, opts) + '}'
            listing_data_supplemental = '{' + '''
                  "partnumber": {},
                  "catalogname": {},
                  "belongstolisting": "0",
                  "sortgroup": 0,
                  "sortgrouptext": "",
                  "paramdesc": "",
                  "showhide": {}
                '''.format('"' + part_number + '"', '"' + brand + '"', {}) + '}'
            payload = '{"partData": {' + '''
                "groupindex": "74",
                "listing_data_essential": {},
                "listing_data_supplemental": {},
                "OptKey": "0-0-0-1"
            '''.format(listing_data_essential, listing_data_supplemental) + '}}'
            payload = payload.replace('\n', '').replace('  ', '')
            rsp_detail.close()

            wb = load_workbook(filename='./Rockauto_basic.xlsx')
            ws = wb['Sheet1']
            serial = ws.max_row
            ws.append([serial, brand, part_number, payload, src, oe, note, url])
            last_row = ws.max_row
            ws.row_dimensions[last_row].height = 20

            for col in range(1, 9):
                cell = ws.cell(row=last_row, column=col)
                cell.font = font
                cell.alignment = alignment
                cell.border = border
            wb.save(filename='./Rockauto_basic.xlsx')
            wb.close()
            if status_code_1 == 200:
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(part_number, test+1, last_row-1, len(Url)))
                break
        except Exception as e:
            # print(f"Error processing URL {url}: {e}")
            time.sleep(0.3)
            if test == num-1:
                print('{}【尝试次数：{}】-error'.format(Part_Number[Url.index(url)], test))
            continue


# jobs = [gevent.spawn(fetch_data, url) for url in Url]
# gevent.joinall(jobs)
pool = pool.Pool(15)
for url in Url:
    pool.spawn(fetch_data, url)
pool.join()
end = time.perf_counter()
print(end-start)

