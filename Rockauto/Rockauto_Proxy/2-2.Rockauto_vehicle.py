# -*- coding: utf-8 -*- 
# @Time ： 2024/7/4 16:04
# @Auth ： Westbroobo
# @File ：2-2.Rockauto_vehicle.py

from gevent import monkey, pool
monkey.patch_all(thread=False)
import time
import requests
from lxml import etree
from tool import UA, Proxy
import json
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
start = time.perf_counter()
df_menu = pd.read_excel('./Rockauto_basic.xlsx', header=0, dtype=str).fillna('')
Brand = df_menu['Brand'].to_list()
Part_Number = df_menu['Part_Number'].to_list()
payloads = df_menu['payload'].to_list()
Pic = df_menu['Pic'].to_list()
OE = df_menu['OE'].to_list()
Note = df_menu['Note'].to_list()
URL =  df_menu['Url'].to_list()

wb = xlsxwriter.Workbook('./Rockauto_detail.xlsx')
ws = wb.add_worksheet('Sheet1')
ws.set_column('A:A', 9)
ws.set_column('B:B', 15)
ws.set_column('C:C', 18)
ws.set_column('D:D', 25)
ws.set_column('E:E', 45)
ws.set_column('F:F', 20)
ws.set_column('G:G', 18)
ws.set_column('H:H', 55)
headings = ['No.', 'Brand', 'Part_Number', 'Vehicle', 'Pic', 'OE', 'Note', 'Url']
head_format = wb.add_format({
    'bold': 1,
    'fg_color': 'cyan',
    'align': 'center',
    'font_name': '等线',
    'size': 10,
    'valign': 'vcenter',
    'border': 1
})
ws.write_row('A1', headings, head_format)
ws.freeze_panes(1, 0)
ws.set_row(0, 20)
wb.close()
font = Font(name='等线', size=9)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))


def fetch_data(payload):
    num = 50
    for test in range(num):
        try:
            data = {
                'func': 'getbuyersguide',
                'payload': payload,
                'api_json_request': '1',
                'sctchecked': '0',
                'scbeenloaded': 'false',
                'curCartGroupID': '',
            }
            headers = UA.get_User_Agent_Requests()
            response = requests.post('https://www.rockauto.com/catalog/catalogapi.php', headers=headers,
                                     data=data, proxies=Proxy.get_Proxy_Requests())
            status_code_2 = response.status_code
            json_data = json.loads(response.text)
            body = json_data['buyersguidepieces']['body']
            tree_2 = etree.HTML(body)
            try:
                trs = tree_2.xpath('./descendant::div[@class="buyersguide-nested"]/div/table/tr')
                content = ''
                for tr in trs:
                    content += tr.xpath('./td[1]/text()')[0] + ' ' + tr.xpath('./td[2]/text()')[0] + ' ' + \
                               tr.xpath('./td[3]/text()')[0] + '\n'
                vehicle = content[:-1]
            except IndexError:
                vehicle = ''
            brand = Brand[payloads.index(payload)]
            part_number = Part_Number[payloads.index(payload)]
            src = Pic[payloads.index(payload)]
            oe = OE[payloads.index(payload)]
            note = Note[payloads.index(payload)]
            url = URL[payloads.index(payload)]
            response.close()
            wb = load_workbook(filename='./Rockauto_detail.xlsx')
            ws = wb['Sheet1']
            serial = ws.max_row
            ws.append([serial, brand, part_number, vehicle, src, oe, note, url])
            last_row = ws.max_row
            ws.row_dimensions[last_row].height = 20

            for col in range(1, 9):
                cell = ws.cell(row=last_row, column=col)
                cell.font = font
                cell.alignment = alignment
                cell.border = border
            wb.save(filename='./Rockauto_detail.xlsx')
            wb.close()
            if status_code_2 == 200:
                print('{}【尝试次数：{}】-success，进度：{}/{}'.
                      format(part_number, test+1, last_row-1, len(Part_Number)))
                break
        except Exception as e:
            # print(f"Error processing URL {url}: {e}")
            time.sleep(0.3)
            if test == num-1:
                print('{}【尝试次数：{}】-error'.format(Part_Number[payloads.index(payload)], test))
            continue


# jobs = [gevent.spawn(fetch_data, payload) for payload in payloads]
# gevent.joinall(jobs)
pool = pool.Pool(15)
for payload in payloads:
    pool.spawn(fetch_data, payload)
pool.join()
wb.close()
end = time.perf_counter()
print(end-start)



